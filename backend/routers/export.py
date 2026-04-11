import csv
import io
from decimal import Decimal
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional

from database import get_pool

router = APIRouter(prefix="/api/export", tags=["export"])

REPORT_QUERIES = {
    "rollup-by-dept": """
        SELECT
            d.dept_name                AS "Отдел",
            COALESCE(SUM(b.plan_rub), 0)  AS "План",
            COALESCE(SUM(b.fact_rub), 0)  AS "Факт",
            COALESCE(SUM(b.plan_rub), 0) - COALESCE(SUM(b.fact_rub), 0) AS "Отклонение",
            COUNT(DISTINCT doc.doc_id)     AS "Кол-во документов"
        FROM department d
        LEFT JOIN budget   b   ON b.dept_id = d.dept_id
        LEFT JOIN document doc ON doc.dept_id = d.dept_id
        GROUP BY d.dept_id, d.dept_name
        ORDER BY d.dept_name
    """,
    "rollup-by-item": """
        SELECT
            bi.item_name               AS "Статья",
            bi.item_category           AS "Категория",
            COALESCE(SUM(b.plan_rub), 0)  AS "План",
            COALESCE(SUM(b.fact_rub), 0)  AS "Факт",
            COALESCE(SUM(b.plan_rub), 0) - COALESCE(SUM(b.fact_rub), 0) AS "Отклонение",
            COUNT(DISTINCT doc.doc_id)     AS "Кол-во документов"
        FROM budget_item bi
        LEFT JOIN budget   b   ON b.item_id = bi.item_id
        LEFT JOIN document doc ON doc.item_id = bi.item_id
        GROUP BY bi.item_id, bi.item_name, bi.item_category
        ORDER BY bi.item_name
    """,
    "rollup-by-quarter": """
        SELECT
            b.budget_year              AS "Год",
            b.budget_quarter           AS "Квартал",
            COALESCE(SUM(b.plan_rub), 0) AS "План",
            COALESCE(SUM(b.fact_rub), 0) AS "Факт",
            COALESCE(SUM(b.plan_rub), 0) - COALESCE(SUM(b.fact_rub), 0) AS "Отклонение"
        FROM budget b
        GROUP BY b.budget_year, b.budget_quarter
        ORDER BY b.budget_year, b.budget_quarter
    """,
    "cross-dept-item": """
        SELECT
            d.dept_name                AS "Отдел",
            bi.item_name               AS "Статья",
            COALESCE(SUM(b.plan_rub), 0) AS "План",
            COALESCE(SUM(b.fact_rub), 0) AS "Факт",
            COALESCE(SUM(b.plan_rub), 0) - COALESCE(SUM(b.fact_rub), 0) AS "Отклонение"
        FROM budget b
        JOIN department  d  ON d.dept_id  = b.dept_id
        JOIN budget_item bi ON bi.item_id = b.item_id
        GROUP BY d.dept_id, d.dept_name, bi.item_id, bi.item_name
        ORDER BY d.dept_name, bi.item_name
    """,
    "slice-by-dept": """
        SELECT
            d.dept_name                AS "Отдел",
            doc.doc_id                 AS "Документ ID",
            dt.type_name               AS "Тип документа",
            bi.item_name               AS "Статья",
            doc.doc_date               AS "Дата",
            doc.doc_amount             AS "Сумма",
            c.contr_name               AS "Контрагент",
            e.last_name || ' ' || e.first_name AS "Ответственный"
        FROM document doc
        LEFT JOIN department  d  ON d.dept_id    = doc.dept_id
        LEFT JOIN doc_type   dt ON dt.type_id    = doc.type_id
        LEFT JOIN budget_item bi ON bi.item_id   = doc.item_id
        LEFT JOIN contractor  c  ON c.contr_inn   = doc.contr_inn
        LEFT JOIN employee   e  ON e.emp_id      = doc.resp_emp_id
        WHERE doc.dept_id = $1
        ORDER BY doc.doc_date DESC
    """,
    "slice-by-contractor": """
        SELECT
            c.contr_name               AS "Контрагент",
            doc.doc_id                 AS "Документ ID",
            d.dept_name                AS "Отдел",
            dt.type_name               AS "Тип документа",
            bi.item_name               AS "Статья",
            doc.doc_date               AS "Дата",
            doc.doc_amount             AS "Сумма",
            e.last_name || ' ' || e.first_name AS "Ответственный"
        FROM document doc
        LEFT JOIN department  d  ON d.dept_id    = doc.dept_id
        LEFT JOIN doc_type   dt ON dt.type_id    = doc.type_id
        LEFT JOIN budget_item bi ON bi.item_id   = doc.item_id
        LEFT JOIN contractor  c  ON c.contr_inn   = doc.contr_inn
        LEFT JOIN employee   e  ON e.emp_id      = doc.resp_emp_id
        WHERE doc.contr_inn = $1
        ORDER BY doc.doc_date DESC
    """,
}

PARAMETERISED_REPORTS = {"slice-by-dept", "slice-by-contractor"}


def _serialize(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


async def _fetch_report(report_type: str, pool, dept_id=None, contr_inn=None):
    query = REPORT_QUERIES.get(report_type)
    if query is None:
        raise HTTPException(404, f"Unknown report type: {report_type}")

    params: list = []
    if report_type == "slice-by-dept":
        if dept_id is None:
            raise HTTPException(400, "dept_id is required for slice-by-dept")
        params.append(dept_id)
    elif report_type == "slice-by-contractor":
        if contr_inn is None:
            raise HTTPException(400, "contr_inn is required for slice-by-contractor")
        params.append(contr_inn)

    async with pool.acquire() as conn:
        rows = await conn.fetch(query, *params)

    return rows


def _rows_to_csv(rows) -> str:
    if not rows:
        return ""
    buf = io.StringIO()
    headers = list(rows[0].keys())
    writer = csv.writer(buf)
    writer.writerow(headers)
    for r in rows:
        writer.writerow([_serialize(r[h]) for h in headers])
    return buf.getvalue()


def _rows_to_xlsx(rows) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    if not rows:
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)

    for row_idx, r in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=_serialize(r[h]))

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


@router.get("/{report_type}")
async def export_report(
    report_type: str,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    dept_id: Optional[int] = None,
    contr_inn: Optional[str] = None,
    pool=Depends(get_pool),
):
    rows = await _fetch_report(report_type, pool, dept_id=dept_id, contr_inn=contr_inn)

    filename = f"{report_type}.{format}"

    if format == "csv":
        content = _rows_to_csv(rows)
        return StreamingResponse(
            io.StringIO(content),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    xlsx_bytes = _rows_to_xlsx(rows)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
